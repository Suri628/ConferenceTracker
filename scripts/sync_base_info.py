"""同步会议基础字段（地点/日期/地区/官网），数据来源：
1. ccfddl/ccf-deadlines 仓库（众包整理自各会议官网，主要数据源）
2. 会议官网直连兜底（ccfddl 里没有这个会议、或没有 TARGET_YEAR 这一届数据时才用）

目标文件 data/Conferences.xlsx，sheet "2027A类会议"。会议列表（数量/名字/顺序）完全以
xlsx 里已有的行为准——本脚本**不会**自动增删或调整会议顺序，也不碰"方向"/"刊物简称"/
"刊物全称"/"出版方"/"URL"（dblp链接，人工维护）这几列，只更新：
- 会议地点（2027）/ 会议时间（2027）/ 所属地区（2027）：ccfddl 匹配上就用 ccfddl 的数据；
  匹配不上或匹配上了但没有 TARGET_YEAR 这届数据，就尝试直接抓"2027 URL"里已有的官网链接，
  从页面正文里抠日期/地点；两边都没有就填 TBD
- 2027 URL：ccfddl 匹配上就用 ccfddl 给的官网链接刷新；匹配不上则保留原值不动（人工维护，
  比如 ISSCC/IEDM/VLSI/WINE 这些不在 CCF 名单里的会议）
- Workshop Submission Deadline（2027）/ 2027 Workshop Proposal URL 这两列不碰，那部分由
  check_workshop_deadlines.py 维护

只有 data/Conferences.xlsx 还不存在时，才会报错要求先手动建好这张表（这个脚本不负责初始化新表）。
"""
import glob
import html
import os
import re
import shutil
import subprocess
import sys
import tempfile

import openpyxl
import requests
import yaml
from openpyxl.styles import Font, PatternFill

REPO_URL = "https://github.com/ccfddl/ccf-deadlines.git"
XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "Conferences.xlsx")
SHEET_NAME = "2027A类会议"

# 表格只关心这一届；到了下一年需要手动把这个数字往前推一年。
TARGET_YEAR = 2027

ACRONYM_COL = "刊物简称"
LOCATION_COL = "会议地点\n（2027）"
DATE_COL = "会议时间\n（2027）"
REGION_COL = "所属地区\n（2027）"
URL_COL = "2027 URL"

HEADERS = {"User-Agent": "Mozilla/5.0 (ConferenceTracker base-info-bot)"}
REQUEST_TIMEOUT = 15

# 本次运行（sync_base_info.py + check_workshop_deadlines.py 算一个完整周期）里有变化的行，
# 整行标黄；每次 sync_base_info.py 开跑会先把上一轮的高亮清空，重新开始记
HIGHLIGHT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)


def set_row_highlight(row, highlighted: bool):
    fill = HIGHLIGHT_FILL if highlighted else NO_FILL
    for cell in row:
        cell.fill = fill

US_STATES = {
    "alabama", "alaska", "arizona", "arkansas", "california", "colorado", "connecticut",
    "delaware", "florida", "georgia", "hawaii", "idaho", "illinois", "indiana", "iowa",
    "kansas", "kentucky", "louisiana", "maine", "maryland", "massachusetts", "michigan",
    "minnesota", "mississippi", "missouri", "montana", "nebraska", "nevada",
    "new hampshire", "new jersey", "new mexico", "new york", "north carolina",
    "north dakota", "ohio", "oklahoma", "oregon", "pennsylvania", "rhode island",
    "south carolina", "south dakota", "tennessee", "texas", "utah", "vermont",
    "virginia", "washington", "west virginia", "wisconsin", "wyoming",
    "al", "ak", "az", "ar", "ca", "co", "ct", "de", "fl", "ga", "hi", "id", "il", "in",
    "ia", "ks", "ky", "la", "me", "md", "ma", "mi", "mn", "ms", "mo", "mt", "ne", "nv",
    "nh", "nj", "nm", "ny", "nc", "nd", "oh", "ok", "or", "pa", "ri", "sc", "sd", "tn",
    "tx", "ut", "vt", "va", "wa", "wv", "wi", "wy",
}

MONTH_NUM = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}

DATE_RANGE_RE = re.compile(
    r"([A-Za-z]+)\.?\s+(\d{1,2})(?:st|nd|rd|th)?"
    r"\s*(?:-|–|—|to)\s*"
    r"(?:([A-Za-z]+)\.?\s+)?(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})"
)

# 所属地区：按"会议地点"里出现的国家名（不分大小写子串匹配，按长度从长到短试，
# 避免"korea"提前命中"south korea"里的子串导致漏掉更精确的匹配）分类。
# 只覆盖这份清单里目前出现过的国家 + 常见补充；遇到没覆盖到的新国家，脚本会保留原有地区值不动
# 并且不会报错，需要时在这里补一条
COUNTRY_REGION = {
    "united states": "北美", "usa": "北美", "u.s.a": "北美", "u.s": "北美", "canada": "北美",
    "mexico": "拉美", "panama": "拉美", "panamá": "拉美", "brazil": "拉美", "argentina": "拉美",
    "chile": "拉美", "colombia": "拉美", "peru": "拉美", "uruguay": "拉美", "costa rica": "拉美",
    "south america": "拉美",
    "united kingdom": "欧洲", "uk": "欧洲", "england": "欧洲", "scotland": "欧洲", "ireland": "欧洲",
    "france": "欧洲", "germany": "欧洲", "netherlands": "欧洲", "denmark": "欧洲", "greece": "欧洲",
    "italy": "欧洲", "spain": "欧洲", "portugal": "欧洲", "sweden": "欧洲", "norway": "欧洲",
    "finland": "欧洲", "switzerland": "欧洲", "austria": "欧洲", "belgium": "欧洲", "poland": "欧洲",
    "czech": "欧洲", "hungary": "欧洲", "iceland": "欧洲", "luxembourg": "欧洲",
    "china": "亚太", "japan": "亚太", "south korea": "亚太", "korea": "亚太", "singapore": "亚太",
    "hong kong": "亚太", "taiwan": "亚太", "vietnam": "亚太", "india": "亚太", "australia": "亚太",
    "new zealand": "亚太", "thailand": "亚太", "malaysia": "亚太", "indonesia": "亚太",
    "philippines": "亚太", "macau": "亚太",
    "morocco": "非洲", "south africa": "非洲", "egypt": "非洲", "nigeria": "非洲", "kenya": "非洲",
    "tunisia": "非洲",
}

# "刊物简称"跟 ccfddl 里的 title 拼写/缩写不完全一致的，手动对齐一下。
# 值可以写 "slug" 或 "slug:CATEGORY"——ccfddl 里有撞名的情况（比如 FSE 既是软工的
# Foundations of Software Engineering 又是密码学的 Fast Software Encryption），
# 不写 CATEGORY 消歧的话，撞名时到底匹配到哪个取决于字典遍历顺序，在 Windows/Linux
# 上可能不一样（本地测试是对的，Actions 上跑出来却是错的，就是这个原因），所以必须显式指定
ALIASES = {
    "usenixatc": "sigopsatc", "fseesec": "fse:SE", "vr": "ieeevr",
    "siggraph": "acmsiggraph", "ubicomp": "ubicompiswc",
    # 不在 ccfddl / CCF 名单里的会议，明确标 None，不去猜；这些走官网直连兜底（见 scrape_official_site）
    "isscc": None, "wine": None, "iedm": None, "vlsi": None,
}

# ccfddl 的 place 字段是"场馆, 城市, 国家"这种三段式时，通用规则分不清楚该丢哪一段
# （跟"城市, 省, 国家"长得一样但处理方式相反），已知的几个手动订正一下
LOCATION_OVERRIDES = {
    "chi": "Pittsburgh, (PA), USA",
}

# ccfddl 的 date 字段有时候没跟着官网更新（众包维护，可能滞后），已知有问题的手动订正一下
# HPCA 2027：ccfddl 写的是 Jan 30-Feb 3，但官网 conf.researchr.org/home/hpca-2027
# 明确写的是 March 20-24（跟同期同地联合举办的 PPoPP 对得上），2026-07 核实过
DATE_OVERRIDES = {
    "hpca": "03.20-03.24",
}


def clone_ccfddl(tmpdir: str) -> str:
    subprocess.run(
        ["git", "clone", "--depth", "1", REPO_URL, tmpdir],
        check=True, capture_output=True,
    )
    return os.path.join(tmpdir, "conference")


def slugify(title: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", title.lower())


def load_ccf_lookup(conference_dir: str) -> dict:
    """(slug, category) -> ccfddl entry，覆盖全部 rank（不只是 CCF-A，比如 IJCAI 是 CCF-B）。
    用 (slug, category) 做 key 而不是单纯 slug，是为了容纳 ccfddl 里撞名的会议——
    两个会议同名但不同 category 时不会互相覆盖，都能查到，靠 ALIASES 里的 CATEGORY 去挑对的那个"""
    lookup = {}
    for path in glob.glob(os.path.join(conference_dir, "*", "*.yml")):
        with open(path, encoding="utf-8") as f:
            docs = yaml.safe_load(f)
        for entry in (docs if isinstance(docs, list) else [docs]):
            if entry:
                lookup[(slugify(entry["title"]), entry.get("sub"))] = entry
    return lookup


def resolve_entry(acronym: str, lookup: dict):
    """按"刊物简称"找 ccfddl 里对应的 entry。撞名的（同一个 slug 对应多个 category）
    必须在 ALIASES 里用 "slug:CATEGORY" 显式指定，不然宁可返回 None 也不瞎猜——
    瞎猜的话结果会取决于字典遍历顺序，不同系统上可能不一样"""
    key = slugify(acronym)
    alias = ALIASES[key] if key in ALIASES else key
    if alias is None:
        return None
    if ":" in alias:
        slug, category = alias.split(":", 1)
        return lookup.get((slug, category))
    matches = [entry for (s, _c), entry in lookup.items() if s == alias]
    return matches[0] if len(matches) == 1 else None


def pick_target_year_conf(confs: list, target_year: int) -> dict:
    for c in confs or []:
        try:
            if int(c.get("year", 0)) == target_year:
                return c
        except (TypeError, ValueError):
            continue
    return {}


def format_date_range(raw: str) -> str:
    """把 ccfddl 里 'July 23-29, 2022' 这类自由文本转成 mm.dd-mm.dd；解析不出来原样返回"""
    if not raw:
        return "TBD"
    m = DATE_RANGE_RE.search(raw)
    if not m:
        return raw
    mon1, day1, mon2, day2, _year = m.groups()
    n1 = MONTH_NUM.get(mon1.lower())
    n2 = MONTH_NUM.get(mon2.lower()) if mon2 else n1
    if not n1 or not n2:
        return raw
    return f"{n1:02d}.{int(day1):02d}-{n2:02d}.{int(day2):02d}"


def format_location(place: str) -> str:
    """'Salt Lake City, Utah' -> 'Salt Lake City, (Utah), USA'；'Kyoto, Japan' 原样两段式保留"""
    if not place:
        return "TBD"
    place = place.rstrip(".").strip()
    parts = [p.strip() for p in place.split(",")]
    if len(parts) < 2:
        return parts[0] if parts else "TBD"
    city, last = ", ".join(parts[:-1]), parts[-1]
    if last.lower() in US_STATES:
        return f"{city}, ({last}), USA"
    if last.upper() in ("USA", "US") and len(parts) >= 3:
        maybe_state = parts[-2]
        if maybe_state.lower() in US_STATES:
            city2 = ", ".join(parts[:-2])
            return f"{city2}, ({maybe_state}), USA"
        return f"{city}, {last}"
    return f"{city}, {last}"


def infer_region(location: str):
    """从地点字符串猜所属地区；猜不出来（没见过的国家/地区名）返回 None，调用方应保留原值不动"""
    if not location or location == "TBD":
        return "TBD"
    lowered = location.lower()
    for country in sorted(COUNTRY_REGION, key=len, reverse=True):
        if country in lowered:
            return COUNTRY_REGION[country]
    return None


# ---- 官网直连兜底：只在 ccfddl 没有这个会议/没有 TARGET_YEAR 这届数据时才用 ----

def fetch(url: str):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 200:
            return resp.text
    except requests.RequestException:
        pass
    return None


def strip_html(raw_html: str) -> str:
    text = re.sub(r"<script.*?</script>", " ", raw_html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<style.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    # 官网页面里常有 &ndash;/&bull; 这类 HTML 实体，不转义的话日期里的连接符会变成字面的
    # "&ndash;" 字符串，正则怎么都匹配不上
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text)


# 地点用"专有名词"（每个词首字母必须大写）拼接，而不是任意字母+空格——否则捕获组会
# 顺着小写单词一路吃进下一句（比如 "...USA next year." 会把 "next year" 也吃进来），
# 或者在遇到句号后继续吃到下一句的开头（比如 "Bangkok, Thailand . News" 里的 "News"）
PROPER_NOUN = r"[A-Z][A-Za-z.\-']*(?:\s[A-Z][A-Za-z.\-']*)*"
LOCATION_HINT_PATTERNS = [
    re.compile(
        r"(?:will be held|will take place|takes place|to be held|held)\b.{0,90}?\bin\s+"
        rf"({PROPER_NOUN},\s*{PROPER_NOUN}(?:,\s*{PROPER_NOUN})?)"
    ),
    re.compile(rf"(?:Venue|Location)\s*[:\-]\s*({PROPER_NOUN},\s*{PROPER_NOUN}(?:,\s*{PROPER_NOUN})?)"),
]


# 官网上常见的另一种写法是"日在前"（跟 ccfddl 自己"月在前"的 'July 23-29, 2022' 约定不同），
# 比如 LICS 官网写的是 'Montreal • 21-24 June 2027'。这两个 pattern 专门补这个格式，
# DAY_FIRST_SAME_MONTH 要求两个日期数字之间不能再出现月份名，避免跟跨月的情形匹配混淆
DAY_FIRST_SAME_MONTH_RE = re.compile(
    r"(\d{1,2})(?:st|nd|rd|th)?\s*(?:-|–|—|to)\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\.?,?\s+(\d{4})"
)
DAY_FIRST_CROSS_MONTH_RE = re.compile(
    r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\.?\s*(?:-|–|—|to)\s*"
    r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\.?,?\s+(\d{4})"
)


def extract_date_from_page(text: str, year: int):
    """页面正文里找日期范围，年份必须匹配 TARGET_YEAR，避免抓到往届的日期。
    先试'月在前'（'March 20-24, 2027'），再试'日在前'（'21-24 June 2027'）两种常见写法"""
    for m in DATE_RANGE_RE.finditer(text):
        mon1, day1, mon2, day2, y = m.groups()
        if int(y) != year:
            continue
        n1 = MONTH_NUM.get(mon1.lower())
        n2 = MONTH_NUM.get(mon2.lower()) if mon2 else n1
        if not n1 or not n2:
            continue
        return f"{n1:02d}.{int(day1):02d}-{n2:02d}.{int(day2):02d}"

    for m in DAY_FIRST_CROSS_MONTH_RE.finditer(text):
        day1, mon1, day2, mon2, y = m.groups()
        if int(y) != year:
            continue
        n1 = MONTH_NUM.get(mon1.lower())
        n2 = MONTH_NUM.get(mon2.lower())
        if not n1 or not n2:
            continue
        return f"{n1:02d}.{int(day1):02d}-{n2:02d}.{int(day2):02d}"

    for m in DAY_FIRST_SAME_MONTH_RE.finditer(text):
        day1, day2, mon, y = m.groups()
        if int(y) != year:
            continue
        n = MONTH_NUM.get(mon.lower())
        if not n:
            continue
        return f"{n:02d}.{int(day1):02d}-{n:02d}.{int(day2):02d}"

    return None


LOCATION_YEAR_WINDOW = 200


def extract_location_from_page(text: str, year: int):
    """地点提取必须在附近（前后 LOCATION_YEAR_WINDOW 字符内）看到 TARGET_YEAR 才采信，
    否则很容易把官网首页上"去年/今年"还没更新的地点当成下一届的地点（比如域名不带年份、
    常年用同一个官网首页的会议，首页可能还停留在上一届的信息）"""
    for pattern in LOCATION_HINT_PATTERNS:
        for m in pattern.finditer(text):
            start, end = m.span()
            window = text[max(0, start - LOCATION_YEAR_WINDOW):end + LOCATION_YEAR_WINDOW]
            if str(year) in window:
                return m.group(1).strip()
    return None


def scrape_official_site(url: str, year: int):
    """返回 (location, date)，任何一项抓不到就是 None；网络失败/无 url 也是 (None, None)"""
    if not url:
        return None, None
    html = fetch(url)
    if not html:
        return None, None
    text = strip_html(html)
    return extract_location_from_page(text, year), extract_date_from_page(text, year)


def guess_next_year_urls(entry, target_year: int):
    """ccfddl 有这个会议但还没收录 TARGET_YEAR 这届时，按"最近一届 link 里的年份"规律猜一个
    候选 URL（比如 pldi26.sigplan.org -> pldi27.sigplan.org，.../sigcomm/2026/ -> .../sigcomm/2027/）。
    只用最近一届的规律猜，不逐年回溯；调用方需要再拿 str(target_year) 是否出现在页面上做验证，
    猜出来的 URL 不保证真实存在"""
    confs = sorted((entry or {}).get("confs") or [], key=lambda c: c.get("year", 0), reverse=True)
    for c in confs:
        year, link = c.get("year"), c.get("link")
        if not link:
            continue
        try:
            year = int(year)
        except (TypeError, ValueError):
            continue
        if year >= target_year:
            continue
        candidates = []
        full_old, full_new = str(year), str(target_year)
        if full_old in link:
            candidates.append(link.replace(full_old, full_new))
        short_old, short_new = f"{year % 100:02d}", f"{target_year % 100:02d}"
        short_pattern = re.compile(r"(?<!\d)" + re.escape(short_old) + r"(?!\d)")
        if short_pattern.search(link):
            candidates.append(short_pattern.sub(short_new, link))
        return list(dict.fromkeys(candidates))
    return []


def verify_guess(url: str, year: int):
    """猜出来的 URL 抓一下，页面正文里得看到 TARGET_YEAR 才算数，避免猜中一个能打开但
    其实是无关页面/占位页的 URL"""
    html = fetch(url)
    if not html:
        return None
    text = strip_html(html)
    if str(year) not in text:
        return None
    return text


HYPERLINK_FONT = Font(color="0563C1", underline="single")


def set_url(cell, url: str):
    """把 url 写进单元格并做成可点击的超链接。
    清空时必须把 cell.hyperlink 也一起清掉——只清 value 的话，openpyxl 存盘再读回来会
    用残留的 hyperlink 把显示文本"复活"成旧链接，哪怕 value 当时明明已经设成空字符串了"""
    cell.value = url or ""
    if url:
        cell.hyperlink = url
        cell.font = HYPERLINK_FONT
    else:
        cell.hyperlink = None
        cell.font = Font()


def refresh_row(row_cells: dict, entry, slug: str):
    """entry 是 ccfddl 的一条会议记录（可能是 None）；刷新 会议地点/会议时间/所属地区/2027 URL"""
    target = pick_target_year_conf((entry or {}).get("confs") or [], TARGET_YEAR)

    if target:
        location = LOCATION_OVERRIDES.get(slug) or format_location(target.get("place", ""))
        date = DATE_OVERRIDES.get(slug) or format_date_range(target.get("date", ""))
        url = target.get("link", "") or row_cells[URL_COL].value or ""
        row_cells[LOCATION_COL].value = location or "TBD"
        row_cells[DATE_COL].value = date or "TBD"
        set_url(row_cells[URL_COL], url)
    else:
        # ccfddl 没有这个会议，或者有会议但还没有 TARGET_YEAR 这届的数据。先试试能不能按
        # 往届 link 的年份规律猜出这一届的官网（猜中就顺带更新"2027 URL"）；猜不中/没有历史
        # 规律可猜，就退化到直接抓"2027 URL"里已经有的官网链接（人工填的，比如 ISSCC/IEDM/
        # VLSI 的官网首页）。抓不到就保留原有值不动——这些行大多是人工订正过的，不能被一次
        # 抓取失败就打回 TBD
        guessed_url, guessed_text = None, None
        for candidate in guess_next_year_urls(entry, TARGET_YEAR):
            text = verify_guess(candidate, TARGET_YEAR)
            if text:
                guessed_url, guessed_text = candidate, text
                break

        if guessed_text:
            scraped_location = extract_location_from_page(guessed_text, TARGET_YEAR)
            scraped_date = extract_date_from_page(guessed_text, TARGET_YEAR)
            set_url(row_cells[URL_COL], guessed_url)
        else:
            existing_url = row_cells[URL_COL].value or ""
            scraped_location, scraped_date = scrape_official_site(existing_url, TARGET_YEAR)

        location = LOCATION_OVERRIDES.get(slug) or scraped_location
        date = DATE_OVERRIDES.get(slug) or scraped_date
        if location:
            row_cells[LOCATION_COL].value = location
        elif not row_cells[LOCATION_COL].value:
            row_cells[LOCATION_COL].value = "TBD"
        if date:
            row_cells[DATE_COL].value = date
        elif not row_cells[DATE_COL].value:
            row_cells[DATE_COL].value = "TBD"

    region = infer_region(row_cells[LOCATION_COL].value)
    if region is not None:
        row_cells[REGION_COL].value = region
    # region is None：地点字符串里没有认识的国家名，保留原有"所属地区"值不动


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"未找到 {XLSX_PATH}，本脚本不负责初始化新表，请先手动建好会议清单。", file=sys.stderr)
        sys.exit(1)

    tmpdir = tempfile.mkdtemp(prefix="ccfddl_")
    try:
        conference_dir = clone_ccfddl(tmpdir)
        lookup = load_ccf_lookup(conference_dir)

        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME]
        header = [c.value for c in ws[1]]
        idx = {name: i for i, name in enumerate(header) if name}

        refreshed, matched, scraped, changed = 0, 0, 0, 0
        for row in ws.iter_rows(min_row=2):
            acronym = row[idx[ACRONYM_COL]].value
            if not acronym:
                continue
            # 新一轮同步开始，先清掉上一轮留下的高亮
            set_row_highlight(row, False)

            entry = resolve_entry(acronym, lookup)
            override_key = slugify(acronym)
            row_cells = {col: row[idx[col]] for col in (LOCATION_COL, DATE_COL, REGION_COL, URL_COL)}
            # openpyxl 存盘再读回来会把 "" 变成 None，比较时要当成一样，否则每次都会被误判成"变了"
            before = {col: (cell.value or "") for col, cell in row_cells.items()}
            refresh_row(row_cells, entry, override_key)
            after = {col: (cell.value or "") for col, cell in row_cells.items()}
            if before != after:
                set_row_highlight(row, True)
                changed += 1
            if entry is not None:
                matched += 1
            else:
                scraped += 1
            refreshed += 1

        wb.save(XLSX_PATH)
        print(f"Refreshed {refreshed} conferences ({changed} changed): {matched} via ccfddl, {scraped} via official-site fallback.")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    main()
