"""抓取/核实每个会议的 workshop proposal 提交截止时间。

目标文件 data/Conferences.xlsx，sheet "2027A类会议"。

策略（由高到低优先级）：
1. 抓 会议URL 页面，找页面里指向 "workshop" 相关的链接并跟进
2. 找不到就用 Brave Search API 兜底搜索 "{acronym} {TARGET_YEAR} workshop proposal deadline"
   （需要环境变量 BRAVE_SEARCH_API_KEY，没配置则跳过这一步，
   结果标记为空并在 备注 里提示需要人工核实）
3. 对拿到的候选页面文本用正则抽取 deadline

用法：
    python scripts/check_workshop_deadlines.py                  # 跑全部会议
    python scripts/check_workshop_deadlines.py --only PPoPP HPCA   # 只跑指定会议（按"刊物简称"，验证用）
"""
import argparse
import os
import re
import sys
from datetime import datetime, timezone

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "Conferences.xlsx")
SHEET_NAME = "2027A类会议"
LOG_SHEET_NAME = "变更日志"
TARGET_YEAR = 2027

ACRONYM_COL = "刊物简称"
URL_COL = "会议URL"
DEADLINE_COL = "Workshop Deadline"
WORKSHOP_URL_COL = "workshopURL"
NOTES_COL = "备注"

# 跟 sync_base_info.py 用同一个颜色；这里只负责"新增"高亮，不清空——
# 清空动作统一放在 sync_base_info.py 开跑的时候做，因为它是每轮周期的第一步
HIGHLIGHT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

HEADERS = {"User-Agent": "Mozilla/5.0 (ConferenceTracker workshop-deadline-bot)"}
REQUEST_TIMEOUT = 15

BRAVE_API_KEY = os.environ.get("BRAVE_SEARCH_API_KEY")

WORKSHOP_LINK_PATTERN = re.compile(r"workshop", re.IGNORECASE)

# 页面文本里匹配"xxx proposal xxx deadline: <日期>"这类描述，日期格式尽量宽松
# 两种常见日期书写顺序：'October 3, 2025' 和 '17 Oct 2025'（会议官网常见 AoE 写法，可能带星期几前缀）
MONTH = (
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?"
    r"|Aug(?:ust)?|Sep(?:t|tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
)
WEEKDAY = r"(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)(?:day)?"
DATE_TOKEN = (
    r"(" + MONTH + r"\.?\s+\d{1,2}[a-z]{0,2},?\s+\d{4}"
    r"|(?:" + WEEKDAY + r"\s+)?\d{1,2}\s+" + MONTH + r"\.?\s+\d{4}"
    r"|\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{4})"
)

# 高精度：要求"workshop"关键词紧邻 deadline，用于还没确认是 workshop 专属页面时（如会议主页）
# "deadline"/"due"都算数（有些官网写"Proposals Due <date>"而不是"Deadline: <date>"）
STRICT_DEADLINE_PATTERNS = [
    re.compile(
        r"workshop[s]?\s*(?:(?:/|and|&)\s*tutorial[s]?)?\s*(?:proposal|paper)?s?\s*"
        r"(?:submission)?\s*(?:deadline|due)\s*[:\-]?\s*" + DATE_TOKEN,
        re.IGNORECASE,
    ),
    re.compile(
        r"(?:call for workshop proposals?|workshop proposal submission)[^.\n]{0,80}" + DATE_TOKEN,
        re.IGNORECASE,
    ),
    # 日期写在"workshop"前面的情况（如 PPoPP 的"Important Dates"表格），窗口收窄到15字符，
    # 避免跨句子把不相关的日期（如 artifact submission deadline）误判成 workshop 的日期
    re.compile(DATE_TOKEN + r"[^.\n]{0,15}workshop[s]?\s*(?:/|and)?\s*tutorial[s]?\s*proposal", re.IGNORECASE),
]

# 宽松：已经确认是通过"workshop"链接跳转进来的页面，不再要求"workshop"字面出现在deadline旁边。
# 这类页面本身就是workshop专属页，"deadline"往往写成"Deadline for submission: <date>"这种
# 没有"proposal"字样紧邻的形式，所以放宽到只要求 deadline 关键词即可，按从严到宽的顺序尝试
LOOSE_DEADLINE_PATTERNS = STRICT_DEADLINE_PATTERNS + [
    re.compile(DATE_TOKEN + r"[^.\n]{0,60}proposal[^.\n]{0,20}deadline", re.IGNORECASE),
    re.compile(r"proposal[^.\n]{0,20}(?:submission)?\s*deadline\s*[:\-]?\s*" + DATE_TOKEN, re.IGNORECASE),
    re.compile(r"deadline\s*(?:for\s+submission)?\s*[:\-]?\s*" + DATE_TOKEN, re.IGNORECASE),
    re.compile(DATE_TOKEN + r"[^.\n]{0,30}deadline", re.IGNORECASE),
]

NOT_APPLICABLE_HINTS = [
    "no workshops are planned",
    "does not host workshops",
    "not accepting workshop proposals",
]


def fetch(url: str):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 200:
            return resp.text
    except requests.RequestException:
        pass
    return None


def find_workshop_link(html: str, base_url: str):
    hrefs = re.findall(r'href=["\']([^"\']+)["\']', html)
    texts_and_hrefs = re.findall(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>([^<]*)</a>', html, re.IGNORECASE)
    candidates = [href for href, text in texts_and_hrefs if WORKSHOP_LINK_PATTERN.search(text)]
    candidates += [href for href in hrefs if WORKSHOP_LINK_PATTERN.search(href)]
    for href in candidates:
        if href.startswith("http"):
            yield href
        elif href.startswith("/"):
            m = re.match(r"(https?://[^/]+)", base_url)
            if m:
                yield m.group(1) + href


def strip_html(html: str) -> str:
    text = re.sub(r"<script.*?</script>", " ", html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<style.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text)


def extract_deadline(text: str, loose: bool = False):
    patterns = LOOSE_DEADLINE_PATTERNS if loose else STRICT_DEADLINE_PATTERNS
    for pattern in patterns:
        m = pattern.search(text)
        if m:
            return m.group(1)
    return None


MONTH_NUM = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}


def normalize_deadline(raw: str) -> str:
    """把抽取出来的各种日期写法统一成 yyyy.mm.dd；解析不出来就原样返回"""
    if not raw:
        return raw
    raw = raw.strip()

    m = re.match(r"(\d{4})-(\d{2})-(\d{2})$", raw)
    if m:
        return f"{m.group(1)}.{m.group(2)}.{m.group(3)}"

    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})$", raw)
    if m:
        month, day, year = m.groups()
        return f"{year}.{int(month):02d}.{int(day):02d}"

    # 'October 3, 2025' / 'Oct. 3rd 2025'
    m = re.match(r"([A-Za-z]+)\.?\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})$", raw)
    if m:
        month_name, day, year = m.groups()
        month = MONTH_NUM.get(month_name.lower())
        if month:
            return f"{year}.{month:02d}.{int(day):02d}"

    # '24 July 2026' / 'Fri 17 Oct 2025'（可能带星期几前缀）
    m = re.match(
        r"(?:" + WEEKDAY + r"\s+)?(\d{1,2})\s+([A-Za-z]+)\.?\s+(\d{4})$", raw
    )
    if m:
        day, month_name, year = m.groups()
        month = MONTH_NUM.get(month_name.lower())
        if month:
            return f"{year}.{month:02d}.{int(day):02d}"

    return raw


def year_looks_stale(deadline_str: str, conf_year) -> bool:
    """抽到的日期年份和记录的会议年份差太多，很可能是抓到了旧年份的页面（官网URL/标题没更新）"""
    if not deadline_str or not conf_year:
        return False
    years = re.findall(r"\d{4}", deadline_str)
    if not years:
        return False
    try:
        conf_year = int(conf_year)
    except (TypeError, ValueError):
        return False
    return all(abs(int(y) - conf_year) > 1 for y in years)


def looks_not_applicable(text: str) -> bool:
    lowered = text.lower()
    return any(hint in lowered for hint in NOT_APPLICABLE_HINTS)


def brave_search(query: str):
    if not BRAVE_API_KEY:
        return []
    try:
        resp = requests.get(
            "https://api.search.brave.com/res/v1/web/search",
            params={"q": query, "count": 5},
            headers={"Accept": "application/json", "X-Subscription-Token": BRAVE_API_KEY},
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        results = resp.json().get("web", {}).get("results", [])
        return [item["url"] for item in results]
    except requests.RequestException as e:
        # 打出来方便在 Actions 日志里排查（比如 key 无效/限流/参数错误），不要静默吞掉
        status = getattr(e.response, "status_code", "?")
        print(f"  [brave_search] request failed ({status}): {e}", file=sys.stderr)
        return []
    except (KeyError, ValueError) as e:
        print(f"  [brave_search] unexpected response format: {e}", file=sys.stderr)
        return []


def check_one(acronym: str, year, official_url: str):
    """返回 (deadline, status, page_url, source_note)"""
    visited = set()
    # (url, html, loose) — loose=True 表示这个页面是顺着"workshop"链接跳转到的，可以放宽抽取条件
    candidate_pages = []

    if official_url:
        html = fetch(official_url)
        if html:
            candidate_pages.append((official_url, html, False))
            for link in find_workshop_link(html, official_url):
                if link not in visited:
                    visited.add(link)
                    sub_html = fetch(link)
                    if sub_html:
                        candidate_pages.append((link, sub_html, True))

    for url, html, loose in candidate_pages:
        text = strip_html(html)
        deadline = extract_deadline(text, loose=loose)
        if deadline:
            if year_looks_stale(deadline, year):
                return None, "not_yet_announced", url, (
                    f"抓到疑似过期页面（页面日期 {deadline} 和会议年份 {year} 对不上），需人工核实：{url}"
                )
            return normalize_deadline(deadline), "confirmed", url, "extracted from official site"
        if looks_not_applicable(text):
            return None, "not_applicable", url, "official site indicates no workshops"

    # 官网猜不中/没抽到，退化到搜索引擎兜底
    query = f"{acronym} {year} workshop proposal deadline call for workshops"
    for link in brave_search(query):
        if link in visited:
            continue
        html = fetch(link)
        if not html:
            continue
        text = strip_html(html)
        deadline = extract_deadline(text, loose=True)
        if deadline:
            if year_looks_stale(deadline, year):
                return None, "not_yet_announced", link, (
                    f"抓到疑似过期页面（页面日期 {deadline} 和会议年份 {year} 对不上），需人工核实：{link}"
                )
            return normalize_deadline(deadline), "confirmed", link, (
                "extracted via search fallback（搜索兜底抽取，未核实页面是否确属该会议，需要人工审核）"
            )

    if not BRAVE_API_KEY:
        return None, "not_yet_announced", "", "需要人工核实官网（未配置搜索API做兜底）"
    return None, "not_yet_announced", "", "抓取+搜索兜底均未找到，需人工核实"


def load_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header) if name}
    return wb, ws, idx


HYPERLINK_FONT = Font(color="0563C1", underline="single")


def set_url(cell, url: str):
    """把 url 写进单元格并做成可点击的超链接。
    清空时必须把 cell.hyperlink 也一起清掉——只清 value 的话，openpyxl 存盘再读回来会
    用残留的 hyperlink 把显示文本"复活"成旧链接，哪怕 value 当时明明已经设成空字符串了"""
    cell.value = url or ""
    if url:
        cell.hyperlink = url
        cell.font = HYPERLINK_FONT
    else:
        cell.hyperlink = None
        cell.font = Font()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--only", nargs="*", help="只处理指定的会议（按\"刊物简称\"，验证用）")
    args = parser.parse_args()

    wb, ws, idx = load_rows(XLSX_PATH)
    if LOG_SHEET_NAME not in wb.sheetnames:
        log_ws = wb.create_sheet(LOG_SHEET_NAME)
        log_ws.append(["刊物简称", "字段", "旧值", "新值", "变更时间"])
    log_ws = wb[LOG_SHEET_NAME]
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    for row in ws.iter_rows(min_row=2):
        acronym = row[idx[ACRONYM_COL]].value
        if not acronym:
            continue
        if args.only and acronym not in args.only:
            continue
        official_url = row[idx[URL_COL]].value

        print(f"Checking {acronym} ...")
        deadline, status, page_url, note = check_one(acronym, TARGET_YEAR, official_url)
        deadline_value = deadline if deadline else "TBD"

        old_deadline = row[idx[DEADLINE_COL]].value
        if deadline_value != old_deadline:
            log_ws.append([acronym, DEADLINE_COL, old_deadline, deadline_value, now])
            for cell in row:
                cell.fill = HIGHLIGHT_FILL

        row[idx[DEADLINE_COL]].value = deadline_value
        set_url(row[idx[WORKSHOP_URL_COL]], page_url)
        row[idx[NOTES_COL]].value = note

    wb.save(XLSX_PATH)
    print(f"Done. Saved {XLSX_PATH}")


if __name__ == "__main__":
    main()
